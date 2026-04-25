"""
S.K. Degree College — Professional Admission Report Generator
Generates a pixel-perfect .xlsx matching the institutional green-gold theme.
"""
import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os

# Try to import Pillow for logo resizing
try:
    from PIL import Image as PILImage
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


# ── Palette ──────────────────────────────────────────────────────────────
DARK_GREEN   = "1B5E20"
MEDIUM_GREEN = "2E7D32"
LIGHT_GREEN  = "E8F5E9"
ROW_ALT      = "F1F8E9"
GOLD         = "F9A825"
WHITE        = "FFFFFF"
GREY         = "757575"
HDR_GREEN    = "388E3C"

# Status badge colors
STATUS_STYLES = {
    "New":       {"bg": "E3F2FD", "fg": "1565C0"},
    "Contacted": {"bg": "FFF3E0", "fg": "E65100"},
    "Confirmed": {"bg": "E8F5E9", "fg": "2E7D32"},
    "Processed": {"bg": "E8F5E9", "fg": "2E7D32"},
    "Rejected":  {"bg": "FFEBEE", "fg": "C62828"},
}

# Column config
HEADERS = [
    "S.No", "Student Name", "Father's Name", "Email", "Phone",
    "Course Applied", "Group", "Status", "Applied Date", "Address"
]
COL_WIDTHS = [5, 22, 20, 28, 14, 30, 12, 12, 22, 40]
TOTAL_COLS = len(HEADERS)  # 10 columns (A-J)

# Thin grey border for data cells
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ── Helpers ──────────────────────────────────────────────────────────────
def fill_row(ws, row, color, start=1, end=TOTAL_COLS):
    """Apply a solid fill to EVERY cell in a row so merged backgrounds render."""
    pf = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(start, end + 1):
        ws.cell(row=row, column=c).fill = pf


def style_header_row(ws, row, text, *, color, font_color=WHITE,
                     size=12, bold=True, italic=False, halign="center"):
    """Merge A:K, paint every cell, write text in column A."""
    last_col = get_column_letter(TOTAL_COLS)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    fill_row(ws, row, color)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = Font(name="Arial", size=size, bold=bold,
                     italic=italic, color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical="center")


# ── Main ─────────────────────────────────────────────────────────────────
def generate_report():
    base = r"c:\Users\janak\Downloads\_Projects\SK degree college"
    input_file  = os.path.join(base, "project docs", "SK_College_Application_Report.xlsx")
    output_file = os.path.join(base, "project docs", "SK_College_Admission_Report_Final.xlsx")
    logo_path   = os.path.join(base, "public", "logo.jpeg")

    # -- 1. Read source data --
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"[ERROR] Cannot read input: {e}")
        return

    rows = []
    for i, r in df.iterrows():
        msg = str(r.get("message", ""))
        gm = re.search(r"Group:\s*([^|]+)", msg)
        am = re.search(r"Address:\s*(.+)", msg)
        group   = gm.group(1).strip() if gm else "N/A"
        address = am.group(1).strip() if am else "N/A"

        created = r.get("created_at")
        try:
            applied = pd.to_datetime(created).strftime("%d/%m/%Y")
        except Exception:
            applied = "N/A"

        rows.append({
            "S.No":            i + 1,
            "Student Name":    r.get("name", "N/A"),
            "Father's Name":   r.get("father_name", "N/A") or "N/A",
            "Email":           r.get("email", "N/A"),
            "Phone":           str(r.get("phone", "N/A")),
            "Course Applied":  r.get("course", "N/A"),
            "Group":           group,
            "Status":          r.get("status", "New") or "New",
            "Applied Date":    applied,
            "Address":         address,
        })

    # ── 2. Create workbook ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Admission Report"

    # Column widths
    for idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ── 3. HEADER SECTION (rows 1-7) ────────────────────────────────────

    # Row 1 — Logo in A1, College Name in B1:J1 (beside logo)
    ws.row_dimensions[1].height = 60
    last_col = get_column_letter(TOTAL_COLS)
    # Fill the entire row dark green
    fill_row(ws, 1, DARK_GREEN)
    # Merge B1:J1 for the college name (A1 is reserved for logo)
    ws.merge_cells(f"B1:{last_col}1")
    title_cell = ws.cell(row=1, column=2)  # B1
    title_cell.value = "S.K. DEGREE COLLEGE & PG COLLEGE"
    title_cell.font = Font(name="Arial", size=20, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — Affiliation (medium green, white text, size 12)
    ws.row_dimensions[2].height = 22
    style_header_row(ws, 2, "Affiliated to Andhra University | Established in 2005",
                     color=MEDIUM_GREEN, font_color=WHITE, size=12, bold=True)

    # Row 3 — Address (no fill, dark green text, size 10)
    ws.row_dimensions[3].height = 18
    last_col = get_column_letter(TOTAL_COLS)
    ws.merge_cells(f"A3:{last_col}3")
    c3 = ws.cell(row=3, column=1)
    c3.value = "School Nagar, Ayyannapet Junction, Vizianagaram, Andhra Pradesh"
    c3.font = Font(name="Arial", size=10, color=DARK_GREEN)
    c3.alignment = Alignment(horizontal="center", vertical="center")

    # Row 4 — Contact (no fill, grey italic, size 9)
    ws.row_dimensions[4].height = 16
    ws.merge_cells(f"A4:{last_col}4")
    c4 = ws.cell(row=4, column=1)
    c4.value = "Ph: 94412 53163 | Email: arunodayaes@yahoo.com | www.skdegreecollege.com"
    c4.font = Font(name="Arial", size=9, italic=True, color=GREY)
    c4.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5 — Gold divider
    ws.row_dimensions[5].height = 8
    ws.merge_cells(f"A5:{last_col}5")
    fill_row(ws, 5, GOLD)

    # Row 6 — Report title (light green bg, dark green text, size 13)
    ws.row_dimensions[6].height = 24
    style_header_row(ws, 6, "ONLINE APPLICATION LIST – ADMISSIONS 2026-27",
                     color=LIGHT_GREEN, font_color=DARK_GREEN, size=13, bold=True)

    # Row 7 — Timestamp (right-aligned, grey italic)
    ws.row_dimensions[7].height = 16
    ws.merge_cells(f"A7:{last_col}7")
    c7 = ws.cell(row=7, column=1)
    now_str = datetime.now().strftime("%d %B %Y | %I:%M %p")
    c7.value = f"Date of Printing: {now_str}"
    c7.font = Font(name="Arial", size=9, italic=True, color=GREY)
    c7.alignment = Alignment(horizontal="right", vertical="center")

    # ── 4. COLUMN HEADERS (row 8) ───────────────────────────────────────
    ws.row_dimensions[8].height = 30
    hdr_fill = PatternFill(start_color=HDR_GREEN, end_color=HDR_GREEN, fill_type="solid")
    hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
    hdr_align = Alignment(horizontal="center", vertical="center")
    hdr_border = Border(
        left=Side(style="thin", color=MEDIUM_GREEN),
        right=Side(style="thin", color=MEDIUM_GREEN),
        top=Side(style="thin", color=MEDIUM_GREEN),
        bottom=Side(style="thin", color=MEDIUM_GREEN),
    )
    for col_num, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = h
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = hdr_border

    # ── 5. DATA ROWS ────────────────────────────────────────────────────
    data_font = Font(name="Arial", size=9)
    alt_fill  = PatternFill(start_color=ROW_ALT, end_color=ROW_ALT, fill_type="solid")

    current_row = 9
    for idx, record in enumerate(rows):
        ws.row_dimensions[current_row].height = 20
        is_even = idx % 2 == 1  # alternate starting from second row

        for col_num, key in enumerate(HEADERS, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = record[key]
            cell.font  = data_font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(key == "Address"),
            )

            # Alternating row fill
            if is_even:
                cell.fill = alt_fill

            # Status badge
            if key == "Status":
                val = str(cell.value)
                style = STATUS_STYLES.get(val)
                if style:
                    cell.fill = PatternFill(start_color=style["bg"],
                                            end_color=style["bg"],
                                            fill_type="solid")
                    cell.font = Font(name="Arial", size=9, bold=True,
                                     color=style["fg"])
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center")

        current_row += 1

    # ── 6. SUMMARY ROW ──────────────────────────────────────────────────
    summary_row = current_row
    ws.row_dimensions[summary_row].height = 25
    ws.merge_cells(f"A{summary_row}:{last_col}{summary_row}")
    fill_row(ws, summary_row, DARK_GREEN)
    sc = ws.cell(row=summary_row, column=1)
    sc.value = f"Total Applications Received: {len(rows)}"
    sc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    sc.alignment = Alignment(horizontal="center", vertical="center")

    # ── 7. FOOTER NOTE ───────────────────────────────────────────────────
    footer_row = summary_row + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    fc = ws.cell(row=footer_row, column=1)
    fc.value = ("This is a computer-generated report from the S.K. Degree "
                "College online admission portal. For queries, contact the "
                "college administration.")
    fc.font = Font(name="Arial", size=8, italic=True, color=GREY)
    fc.alignment = Alignment(horizontal="center", vertical="center")

    # ── 8. LOGO ──────────────────────────────────────────────────────────
    if os.path.exists(logo_path):
        try:
            if HAS_PILLOW:
                pil = PILImage.open(logo_path)
                ratio = 75 / pil.height
                new_w = int(pil.width * ratio)
                pil = pil.resize((new_w, 75))
                temp = os.path.join(base, "scratch", "_temp_logo.png")
                pil.save(temp)
                img = XLImage(temp)
            else:
                img = XLImage(logo_path)
                img.height = 75
                img.width  = 75
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"[WARN] Logo insertion skipped: {e}")

    # ── 9. PRINT & FREEZE SETTINGS ──────────────────────────────────────
    ws.freeze_panes = "A9"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:8"

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(output_file)

    # Cleanup temp logo
    temp = os.path.join(base, "scratch", "_temp_logo.png")
    if os.path.exists(temp):
        os.remove(temp)

    print(f"[OK] Report saved -> {output_file}")


if __name__ == "__main__":
    generate_report()
